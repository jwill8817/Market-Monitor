"""
Academic long/short factor data.

Primary source: Kenneth French Data Library (Dartmouth) — free, no key.
  - Daily US factors:   Mkt-RF, SMB, HML, RMW, CMA, Mom, ST_Rev, LT_Rev
  - Monthly US factors: same set, longer history
  - Monthly regional:   Developed / Emerging 5-factor

Returns are L/S (dollar-neutral) portfolio returns in percent on FRED-style
download; we store them as DECIMAL periodic returns and compound for windows.

Other credible sources (AQR, Open Source Asset Pricing, JKP) are monthly and
either require registration (JKP) or very large downloads (OSAP); hooks are
left for future addition.
"""
import urllib.request
import io
import zipfile
import datetime
import statistics

_KF_BASE = "https://mba.tuck.dartmouth.edu/pages/faculty/ken.french/ftp/"
_UA = "JAWS/1.0 jwill8817@gmail.com"
_TIMEOUT = 25

# (display_name, kf_zip_file, column_in_file)
KF_DAILY = [
    ("Market (Mkt-RF)", "F-F_Research_Data_5_Factors_2x3_daily_CSV.zip", "Mkt-RF"),
    ("Size (SMB)",       "F-F_Research_Data_5_Factors_2x3_daily_CSV.zip", "SMB"),
    ("Value (HML)",      "F-F_Research_Data_5_Factors_2x3_daily_CSV.zip", "HML"),
    ("Profitability (RMW)","F-F_Research_Data_5_Factors_2x3_daily_CSV.zip","RMW"),
    ("Investment (CMA)", "F-F_Research_Data_5_Factors_2x3_daily_CSV.zip", "CMA"),
    ("Momentum (Mom)",   "F-F_Momentum_Factor_daily_CSV.zip",            "Mom"),
    ("Short-Term Rev",   "F-F_ST_Reversal_Factor_daily_CSV.zip",         "ST_Rev"),
    ("Long-Term Rev",    "F-F_LT_Reversal_Factor_daily_CSV.zip",         "LT_Rev"),
]

KF_MONTHLY = [
    ("Market (Mkt-RF)", "F-F_Research_Data_5_Factors_2x3_CSV.zip", "Mkt-RF"),
    ("Size (SMB)",       "F-F_Research_Data_5_Factors_2x3_CSV.zip", "SMB"),
    ("Value (HML)",      "F-F_Research_Data_5_Factors_2x3_CSV.zip", "HML"),
    ("Profitability (RMW)","F-F_Research_Data_5_Factors_2x3_CSV.zip","RMW"),
    ("Investment (CMA)", "F-F_Research_Data_5_Factors_2x3_CSV.zip", "CMA"),
    ("Momentum (Mom)",   "F-F_Momentum_Factor_CSV.zip",            "Mom"),
    ("Short-Term Rev",   "F-F_ST_Reversal_Factor_CSV.zip",         "ST_Rev"),
    ("Long-Term Rev",    "F-F_LT_Reversal_Factor_CSV.zip",         "LT_Rev"),
    ("Dev Market",       "Developed_5_Factors_CSV.zip",            "Mkt-RF"),
    ("Dev Value (HML)",  "Developed_5_Factors_CSV.zip",            "HML"),
    ("Dev Momentum",     "Developed_Mom_Factor_CSV.zip",           "WML"),
    ("EM Market",        "Emerging_5_Factors_CSV.zip",             "Mkt-RF"),
    ("EM Value (HML)",   "Emerging_5_Factors_CSV.zip",             "HML"),
]

# In-process cache: filename -> {column: [(date, decimal_return), ...]}
_CACHE = {}

# ── AQR monthly factors (Excel downloads) ───────────────────
# (display_name, url, country_column, sign)
# sign=-1 flips AQR's "low minus high" BAB into "high minus low beta".
_AQR_BASE = "https://www.aqr.com/-/media/AQR/Documents/Insights/Data-Sets/"
AQR_MONTHLY = [
    ("Quality H-L (QMJ)", _AQR_BASE + "Quality-Minus-Junk-Factors-Monthly.xlsx",
     "USA",  1, "AQR Quality-Minus-Junk (high quality − junk), US"),
    ("Betting-Against-Beta", _AQR_BASE + "Betting-Against-Beta-Equity-Factors-Monthly.xlsx",
     "USA", 1, "AQR Betting-Against-Beta (low-beta − high-beta), US"),
]
_AQR_CACHE = {}   # (url, column) -> [(date, decimal)]

# ── Canonical factor definitions (for exports / documentation) ──────────────
_KF_URL = "https://mba.tuck.dartmouth.edu/pages/faculty/ken.french/data_library.html"
_AQR_URL = "https://www.aqr.com/Insights/Datasets"
# name -> {source, provider_url, long_leg, short_leg, definition, construction}
FACTOR_DEFS = {
    "Market (Mkt-RF)": {
        "source": "Fama-French (Ken French Data Library)", "url": _KF_URL,
        "long": "Value-weighted US equity market", "short": "1-month US Treasury bill",
        "definition": "The value-weighted return of the entire US stock market in excess of the "
                      "risk-free rate — i.e. the broad equity risk premium (not a long/short pairing).",
        "construction": "Return of the CRSP value-weighted portfolio of all US common stocks minus the "
                        "1-month Treasury-bill rate."},
    "Size (SMB)": {
        "source": "Fama-French (Ken French Data Library)", "url": _KF_URL,
        "long": "Small-cap stocks", "short": "Large-cap stocks",
        "definition": "Small Minus Big: a size-tilted pairing that goes long small-cap stocks and short "
                      "large-cap stocks, holding value exposure neutral.",
        "construction": "Average return of the three small-cap portfolios minus the three large-cap "
                        "portfolios, from independent 2×3 sorts on size and book-to-market."},
    "Value (HML)": {
        "source": "Fama-French (Ken French Data Library)", "url": _KF_URL,
        "long": "Cheap 'value' stocks (high book-to-market)", "short": "Expensive 'growth' stocks (low book-to-market)",
        "definition": "High Minus Low: a size-neutral pairing that goes long cheap (high book-to-market) "
                      "stocks and short expensive (low book-to-market) stocks.",
        "construction": "Average return of the two high book-to-market portfolios minus the two low "
                        "book-to-market portfolios, from 2×3 sorts on size and book-to-market."},
    "Profitability (RMW)": {
        "source": "Fama-French (Ken French Data Library)", "url": _KF_URL,
        "long": "Robust (high) operating-profitability firms", "short": "Weak (low) operating-profitability firms",
        "definition": "Robust Minus Weak: a size-neutral pairing that goes long firms with robust operating "
                      "profitability and short firms with weak profitability.",
        "construction": "Average return of the two robust-profitability portfolios minus the two weak-"
                        "profitability portfolios, from 2×3 sorts on size and operating profitability."},
    "Investment (CMA)": {
        "source": "Fama-French (Ken French Data Library)", "url": _KF_URL,
        "long": "Conservative firms (low asset growth)", "short": "Aggressive firms (high asset growth)",
        "definition": "Conservative Minus Aggressive: a size-neutral pairing that goes long firms that "
                      "invest conservatively (grow assets slowly) and short firms that invest aggressively.",
        "construction": "Average return of the two low-investment portfolios minus the two high-investment "
                        "portfolios, from 2×3 sorts on size and prior-year asset growth."},
    "Momentum (Mom)": {
        "source": "Fama-French / Carhart (Ken French Data Library)", "url": _KF_URL,
        "long": "Recent winners (high prior 2–12m return)", "short": "Recent losers (low prior 2–12m return)",
        "definition": "A size-neutral pairing that goes long recent winners and short recent losers, "
                      "ranked on the prior 2–12 month return (skipping the most recent month).",
        "construction": "Average return of the two high-prior-return portfolios minus the two low-prior-"
                        "return portfolios, from 2×3 sorts on size and prior (2–12 month) return."},
    "Short-Term Rev": {
        "source": "Fama-French (Ken French Data Library)", "url": _KF_URL,
        "long": "Last month's losers", "short": "Last month's winners",
        "definition": "A size-neutral pairing that goes long last month's losers and short last month's "
                      "winners — the one-month reversal effect.",
        "construction": "Average return of the two low-prior-month portfolios minus the two high-prior-"
                        "month portfolios, from 2×3 sorts on size and prior 1-month return."},
    "Long-Term Rev": {
        "source": "Fama-French (Ken French Data Library)", "url": _KF_URL,
        "long": "3–5 year losers (months 13–60)", "short": "3–5 year winners (months 13–60)",
        "definition": "A size-neutral pairing that goes long multi-year losers and short multi-year "
                      "winners, ranked on the return over months 13–60 — the long-horizon reversal effect.",
        "construction": "Average return of the two low-prior portfolios minus the two high-prior "
                        "portfolios, from 2×3 sorts on size and return over months 13–60."},
    "Quality H-L (QMJ)": {
        "source": "AQR Capital Management", "url": _AQR_URL,
        "long": "High-quality firms (profitable, growing, safe, high payout)", "short": "Low-quality 'junk' firms",
        "definition": "Quality Minus Junk: a market-neutral pairing that goes long high-quality firms "
                      "(profitable, growing, safe, high-payout) and short low-quality 'junk' firms.",
        "construction": "Within size groups, rank on a composite quality z-score (profitability, growth, "
                        "safety, payout); long the top, short the bottom, dollar- and beta-neutral. US series."},
    "Betting-Against-Beta": {
        "source": "AQR Capital Management", "url": _AQR_URL,
        "long": "Low-beta stocks (leveraged up to β=1)", "short": "High-beta stocks (de-leveraged to β=1)",
        "definition": "Betting Against Beta: a market-neutral pairing that goes long a basket of low-beta "
                      "stocks (leveraged up) and short a basket of high-beta stocks (de-leveraged), so the "
                      "combined portfolio has ~zero market beta.",
        "construction": "Rank by estimated market beta; long low-beta levered to β=1, short high-beta "
                        "de-levered to β=1, so the pair is beta-neutral by construction. US series."},
}
# regional variants share the base definition, tagged by region
_REGION_TAG = {"Dev ": "Developed ex-US universe", "EM ": "Emerging-markets universe"}
_REGION_BASE = {"Dev Market": "Market (Mkt-RF)", "Dev Value (HML)": "Value (HML)",
                "Dev Momentum": "Momentum (Mom)", "EM Market": "Market (Mkt-RF)",
                "EM Value (HML)": "Value (HML)"}


# Actual download source we pull each series from (built from the KF/AQR tables above).
_KF_FILE = {}   # base name -> {"M": zipfile, "D": zipfile}
for _nm, _f, _c in KF_MONTHLY:
    _KF_FILE.setdefault(_nm, {})["M"] = _f
for _nm, _f, _c in KF_DAILY:
    _KF_FILE.setdefault(_nm, {})["D"] = _f
_AQR_FILE = {r[0]: r[1] for r in AQR_MONTHLY}   # base name -> full xlsx url


def _pull_source(base, freq):
    """(human-readable source, exact file URL) we actually download `base` from."""
    if base in _AQR_FILE:
        url = _AQR_FILE[base]
        return f"AQR Data Library — {url.rsplit('/', 1)[-1]}", url
    if base in _KF_FILE:
        files = _KF_FILE[base]
        f = files.get(freq) or files.get("M") or next(iter(files.values()))
        return f"Ken French Data Library — {f}", _KF_BASE + f
    return "", ""


def factor_definition(name):
    """Return the definition dict for a factor display name (with or without a
    ' (M)'/' (D)' frequency suffix, and handling Dev/EM regional variants).
    Includes the exact file we pull the series from. Returns None if the name
    isn't a recognized academic factor."""
    base = name.strip(); freq = "M"
    for tag in (" [factor]", " [macro]"):        # picker labels carry a source tag
        if base.endswith(tag):
            base = base[:-len(tag)].strip()
    for suf, fq in ((" (M)", "M"), (" (D)", "D")):
        if base.endswith(suf):
            base = base[:-len(suf)].strip(); freq = fq
    if base in FACTOR_DEFS:
        src, url = _pull_source(base, freq)
        return {**FACTOR_DEFS[base], "factor": base, "region": "US",
                "pull_source": src, "pull_url": url}
    if base in _REGION_BASE:
        d = dict(FACTOR_DEFS[_REGION_BASE[base]])
        region = next((v for k, v in _REGION_TAG.items() if base.startswith(k)), "US")
        d["definition"] = f"{d['definition']} ({region}.)"
        src, url = _pull_source(base, freq)   # regional zips are keyed by the regional name
        return {**d, "factor": base, "region": region, "pull_source": src, "pull_url": url}
    return None


def _fetch_aqr_series(url, column, sign=1):
    """Download an AQR monthly factor Excel and return [(date, decimal_return)]."""
    key = (url, column)
    if key in _AQR_CACHE:
        base = _AQR_CACHE[key]
    else:
        import openpyxl, io
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=40) as r:
            raw = r.read()
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]          # first sheet = "<FACTOR> Factors"
        rows = list(ws.iter_rows(values_only=True))
        hdr_i = next((i for i, rr in enumerate(rows)
                      if rr and rr[0] == "DATE"), None)
        base = []
        if hdr_i is not None:
            hdr = rows[hdr_i]
            if column in hdr:
                ci = hdr.index(column)
                for rr in rows[hdr_i + 1:]:
                    if ci >= len(rr):
                        continue
                    d, v = rr[0], rr[ci]
                    if d is None or not isinstance(v, (int, float)):
                        continue
                    if isinstance(v, bool):
                        continue
                    if isinstance(d, datetime.datetime):
                        dd = d.date()
                    elif isinstance(d, datetime.date):
                        dd = d
                    else:
                        try:
                            dd = datetime.datetime.strptime(str(d), "%m/%d/%Y").date()
                        except Exception:
                            continue
                    base.append((dd, float(v)))
        _AQR_CACHE[key] = base
    return [(d, v * sign) for d, v in base]


def _download_zip_csv(filename):
    """Download a Ken French zip and return the contained CSV text."""
    url = _KF_BASE + filename
    req = urllib.request.Request(url, headers={"User-Agent": _UA})
    with urllib.request.urlopen(req, timeout=_TIMEOUT) as r:
        raw = r.read()
    zf = zipfile.ZipFile(io.BytesIO(raw))
    name = zf.namelist()[0]
    return zf.read(name).decode("latin-1")


def _parse_kf_csv(text):
    """Parse a Ken French CSV → {column: [(date, decimal_return)]}.
    Stops at the first non-data block (e.g. the annual section)."""
    lines = text.splitlines()
    header = None
    start = None
    for i, ln in enumerate(lines):
        s = ln.strip()
        if s.startswith(",") and any(c.isalpha() for c in s):
            header = [h.strip() for h in ln.split(",")]
            start = i + 1
            break
    if header is None:
        return {}
    cols = [c for c in header[1:] if c]   # drop leading empty cell

    dates, rows = [], []
    for ln in lines[start:]:
        parts = [p.strip() for p in ln.split(",")]
        tok = parts[0]
        if not tok or not tok.isdigit() or len(tok) not in (6, 8):
            break   # blank line or start of annual block → stop
        try:
            if len(tok) == 8:
                d = datetime.date(int(tok[:4]), int(tok[4:6]), int(tok[6:8]))
            else:
                d = datetime.date(int(tok[:4]), int(tok[4:6]), 1)
            vals = [float(x) for x in parts[1:1 + len(cols)]]
        except Exception:
            break
        dates.append(d)
        rows.append(vals)

    out = {}
    for ci, c in enumerate(cols):
        series = []
        for ri in range(len(dates)):
            v = rows[ri][ci]
            if v <= -99:        # KF missing-data sentinel
                continue
            series.append((dates[ri], v / 100.0))   # percent → decimal
        out[c] = series
    return out


def _get_file(filename):
    if filename not in _CACHE:
        _CACHE[filename] = _parse_kf_csv(_download_zip_csv(filename))
    return _CACHE[filename]


# ── Window return math ──────────────────────────────────────

def _cum_return(series, start_date=None, end_date=None):
    """Compound decimal returns in (start_date, end_date] → percent total."""
    prod = 1.0
    n = 0
    for d, r in series:
        if start_date is not None and d < start_date:
            continue
        if end_date is not None and d > end_date:
            continue
        prod *= (1.0 + r)
        n += 1
    return (prod - 1.0) * 100.0 if n else None


def _quarter_start(today):
    q_first_month = 3 * ((today.month - 1) // 3) + 1
    return datetime.date(today.year, q_first_month, 1)


def _window_returns(series, is_daily):
    """Compute the standard monitor windows for one factor series (percent)."""
    if not series:
        return {}
    today = series[-1][0]
    last_n = lambda n: _cum_return(series[-n:]) if len(series) >= 1 else None

    out = {}
    if is_daily:
        out["1D"] = series[-1][1] * 100.0
        out["1W"] = _cum_return(series[-5:])
        out["1M"] = _cum_return(series[-21:])
        out["3M"] = _cum_return(series[-63:])
        yr = 252
    else:
        out["1D"] = None
        out["1W"] = None
        out["1M"] = series[-1][1] * 100.0
        out["3M"] = _cum_return(series[-3:])
        yr = 12

    out["MTD"] = _cum_return(series, datetime.date(today.year, today.month, 1))
    out["QTD"] = _cum_return(series, _quarter_start(today))
    out["YTD"] = _cum_return(series, datetime.date(today.year, 1, 1))
    out["1Y"]  = _cum_return(series[-yr:])     if len(series) >= 2 else None
    out["3Y"]  = _cum_return(series[-yr*3:])
    out["5Y"]  = _cum_return(series[-yr*5:])
    out["7Y"]  = _cum_return(series[-yr*7:])
    out["10Y"] = _cum_return(series[-yr*10:])
    return out


def build_factor_row(name, series, is_daily, custom_start=None, custom_end=None):
    win = _window_returns(series, is_daily)
    if custom_start is not None:
        win["Custom"] = _cum_return(series, custom_start, custom_end)
    else:
        win["Custom"] = None

    # All-time min/max single-period return + annualized full-period stats
    rets = [r for _, r in series]
    out = {
        "name":       name,
        "is_daily":   is_daily,
        "windows":    win,
        "start":      str(series[0][0]) if series else "",
        "end":        str(series[-1][0]) if series else "",
        "n_obs":      len(series),
        "raw_dates":  [d for d, _ in series],
        "raw_rets":   rets,
    }
    return out


def fetch_factors(custom_start=None, custom_end=None, which="both"):
    """
    Returns dict with 'daily' and 'monthly' lists of factor rows.
    custom_start / custom_end: datetime.date for the Custom column.
    """
    result = {"daily": [], "monthly": []}

    if which in ("both", "daily"):
        for name, fn, col in KF_DAILY:
            try:
                data = _get_file(fn)
                series = data.get(col, [])
                if series:
                    result["daily"].append(
                        build_factor_row(name, series, True, custom_start, custom_end))
            except Exception as e:
                result["daily"].append({"name": name, "error": str(e), "is_daily": True})

    if which in ("both", "monthly"):
        for name, fn, col in KF_MONTHLY:
            try:
                data = _get_file(fn)
                series = data.get(col, [])
                if series:
                    result["monthly"].append(
                        build_factor_row(name, series, False, custom_start, custom_end))
            except Exception as e:
                result["monthly"].append({"name": name, "error": str(e), "is_daily": False})

        # AQR monthly factors (Quality-Minus-Junk, High−Low Beta)
        for name, url, col, sign, _desc in AQR_MONTHLY:
            try:
                series = _fetch_aqr_series(url, col, sign)
                if series:
                    result["monthly"].append(
                        build_factor_row(name, series, False, custom_start, custom_end))
            except Exception as e:
                result["monthly"].append({"name": name, "error": str(e), "is_daily": False})

    return result


def cumulative_series(raw_dates, raw_rets, start_date=None):
    """Return (dates, cumulative_pct) growth-of-1 series for charting."""
    ds, vs = [], []
    prod = 1.0
    started = start_date is None
    for d, r in zip(raw_dates, raw_rets):
        if start_date is not None and d < start_date:
            continue
        prod *= (1.0 + r)
        ds.append(d)
        vs.append((prod - 1.0) * 100.0)
    return ds, vs
