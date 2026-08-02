"""
SIFMA issuance parsing (used by the scheduled refresh agent, not the live app).

SIFMA blocks automated scraping and JS-injects its detailed-file download links, so the
deployed app never fetches SIFMA directly — it reads committed CSVs. This module provides
robust parsers so the scheduled cloud agent (or a human running it locally) can turn a SIFMA
workbook into our tidy schema.

- parse_fixed_income_issuance(raw_xlsx_bytes): the aggregate "US Fixed Income Securities
  Statistics" file → annual issuance by broad asset class (UST/MBS/Corporates/Munis/Agency/ABS).
- parse_monthly_sheet(raw_xlsx_bytes, ...): generic monthly-issuance sheet → tidy month rows
  (used for the detailed Corporate / Equity files once their bytes are provided via upload).

Both take FILE BYTES so they work identically on a live download or an uploaded file.
"""
import io
import datetime


def _wb(raw):
    import openpyxl
    return openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)


def parse_fixed_income_issuance(raw):
    """Aggregate file's 'Issuance' sheet → [{period, asset_class, value_bn}] (annual $bn)."""
    wb = _wb(raw)
    if "Issuance" not in wb.sheetnames:
        return []
    rows = list(wb["Issuance"].iter_rows(values_only=True))
    hdr_i = next((i for i, r in enumerate(rows)
                  if r and any(str(c).strip() == "UST" for c in r if c is not None)), None)
    if hdr_i is None:
        return []
    hdr = rows[hdr_i]
    cols = {}                                   # column index -> asset-class label
    for j, c in enumerate(hdr):
        lab = str(c).strip() if c is not None else ""
        if lab in ("UST", "MBS", "Corporates", "Munis", "Agency", "ABS"):
            cols[j] = lab
    out = []
    for r in rows[hdr_i + 1:]:
        if not r or not isinstance(r[0], (int, float)):
            continue
        yr = int(r[0])
        if not (1980 <= yr <= 2100):
            continue
        for j, lab in cols.items():
            v = r[j] if j < len(r) else None
            if isinstance(v, (int, float)):
                out.append({"period": yr, "asset_class": lab, "value_bn": round(float(v), 2)})
    return out


def parse_monthly_sheet(raw, sheet=None, asset_class="", category_map=None):
    """Generic monthly-issuance parser: finds a date column + numeric category columns.
    category_map: {source_column_header: our_category}. Returns [{month, asset_class,
    category, value_bn}]. Kept permissive because SIFMA detail layouts vary by file."""
    wb = _wb(raw)
    name = sheet or ("Issuance" if "Issuance" in wb.sheetnames else wb.sheetnames[0])
    rows = list(wb[name].iter_rows(values_only=True))
    # header = first row containing any mapped column name
    cmap = category_map or {}
    hdr_i = None
    for i, r in enumerate(rows):
        labs = [str(c).strip() if c is not None else "" for c in (r or [])]
        if any(h in labs for h in cmap):
            hdr_i = i; hdr = labs; break
    if hdr_i is None:
        return []
    colidx = {hdr.index(h): cat for h, cat in cmap.items() if h in hdr}
    out = []
    for r in rows[hdr_i + 1:]:
        if not r:
            continue
        d = _as_month(r[0])
        if d is None:
            continue
        for j, cat in colidx.items():
            v = r[j] if j < len(r) else None
            if isinstance(v, (int, float)):
                out.append({"month": d.isoformat(), "asset_class": asset_class,
                            "category": cat, "value_bn": round(float(v), 2)})
    return out


def _as_month(v):
    """Coerce a cell to a month-end date, or None."""
    if isinstance(v, datetime.datetime):
        v = v.date()
    if isinstance(v, datetime.date):
        return _month_end(v)
    s = str(v).strip() if v is not None else ""
    for fmt in ("%b-%y", "%b %Y", "%B %Y", "%m/%d/%Y", "%Y-%m", "%Y-%m-%d"):
        try:
            return _month_end(datetime.datetime.strptime(s, fmt).date())
        except Exception:
            continue
    return None


def _month_end(d):
    import calendar
    return datetime.date(d.year, d.month, calendar.monthrange(d.year, d.month)[1])
